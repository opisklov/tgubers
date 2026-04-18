import argparse
import json
from collections import defaultdict
from datetime import datetime
from itertools import combinations
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, numbers
from tqdm import tqdm

from classifiers import get_classifier, BaseClassifier


def get_media_type(media: dict | None) -> str:
    """Extract media type from media object."""
    if media is None:
        return "text"

    media_class = media.get("_", "text")

    if media_class == "MessageMediaPhoto":
        return "photo"
    elif media_class == "MessageMediaDocument":
        doc = media.get("document", {})
        attributes = doc.get("attributes", [])

        for attr in attributes:
            attr_type = attr.get("_", "")
            if attr_type == "DocumentAttributeAudio":
                if attr.get("voice"):
                    return "voice"
                return "audio"
            elif attr_type == "DocumentAttributeVideo":
                if attr.get("round_message"):
                    return "video_note"
                return "video"
            elif attr_type == "DocumentAttributeSticker":
                return "sticker"
            elif attr_type == "DocumentAttributeAnimated":
                return "animation"

        return "document"
    elif media_class == "MessageMediaWebPage":
        return "webpage"
    elif media_class == "MessageMediaPoll":
        return "poll"
    elif media_class == "MessageMediaGeo":
        return "geo"
    elif media_class == "MessageMediaContact":
        return "contact"

    return media_class.replace("MessageMedia", "").lower()


def format_reactions(reactions: dict | None) -> str:
    """Format reactions as emoji: count pairs."""
    if reactions is None:
        return ""

    results = reactions.get("results", [])
    if not results:
        return ""

    reaction_parts = []
    for r in results:
        reaction = r.get("reaction", {})
        emoticon = reaction.get("emoticon", "")
        count = r.get("count", 0)
        if emoticon:
            reaction_parts.append(f"{emoticon}: {count}")

    return ", ".join(reaction_parts)


POSITIVE_REACTIONS = {
    "👍", "❤", "🔥", "👏", "🥰", "❤‍🔥", "💯", "⚡", "🎉", "🤩", "😍",
    "🏆", "👌", "🆒", "😊", "🥳", "✨", "🫡", "💪", "🙏", "🎊", "😘", "🤗",
}
NEGATIVE_REACTIONS = {"👎", "💩", "🤮", "😡", "🤬", "😤", "🖕", "😢"}


def total_reactions(reactions: dict | None) -> int:
    if reactions is None:
        return 0
    return sum(r.get("count", 0) for r in reactions.get("results", []))


def split_reactions(reactions: dict | None) -> tuple[int, int, int]:
    """Return (positive, negative, neutral) reaction counts."""
    if reactions is None:
        return 0, 0, 0
    pos = neg = neu = 0
    for r in reactions.get("results", []):
        emoticon = r.get("reaction", {}).get("emoticon", "")
        count = r.get("count", 0)
        if emoticon in POSITIVE_REACTIONS:
            pos += count
        elif emoticon in NEGATIVE_REACTIONS:
            neg += count
        else:
            neu += count
    return pos, neg, neu


def parse_datetime(date_str: str | None) -> tuple[str, str]:
    """Parse ISO datetime string to date and time strings."""
    if not date_str:
        return "", ""

    try:
        dt = datetime.fromisoformat(date_str)
        return dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M:%S")
    except ValueError:
        return "", ""


def post_quarter(date_str: str | None) -> str | None:
    """Return quarter string like '2024-Q1', or None if date is missing/invalid."""
    if not date_str:
        return None
    try:
        dt = datetime.fromisoformat(date_str)
        return f"{dt.year}-Q{(dt.month - 1) // 3 + 1}"
    except ValueError:
        return None


def _empty_bucket() -> dict:
    return {"primary": 0, "total": 0,
            "reactions": 0, "pos": 0, "neg": 0, "neu": 0,
            "views": 0, "view_posts": 0}


def compute_statistics(posts: list[dict]) -> dict:
    """Compute theme statistics across all posts."""
    primary_count: dict[str, int] = defaultdict(int)   # theme is theme_1
    any_count: dict[str, int] = defaultdict(int)        # theme appears anywhere
    reaction_totals: dict[str, int] = defaultdict(int)
    pos_totals: dict[str, int] = defaultdict(int)
    neg_totals: dict[str, int] = defaultdict(int)
    neu_totals: dict[str, int] = defaultdict(int)
    view_totals: dict[str, int] = defaultdict(int)
    post_counts: dict[str, int] = defaultdict(int)      # posts with views data
    cooccurrence: dict[tuple[str, str], int] = defaultdict(int)

    # quarterly[(theme, "2024-Q1")] → stats bucket
    quarterly: dict[tuple[str, str], dict] = defaultdict(_empty_bucket)
    # total classified posts per quarter (for % columns)
    quarter_totals: dict[str, int] = defaultdict(int)

    for post in posts:
        themes = post.get("themes") or []
        if not themes:
            continue

        raw_reactions = post.get("reactions")
        reactions = total_reactions(raw_reactions)
        pos, neg, neu = split_reactions(raw_reactions)
        views = post.get("views") or 0
        quarter = post_quarter(post.get("date"))

        primary_count[themes[0]] += 1
        if quarter:
            quarter_totals[quarter] += 1

        for theme in themes:
            any_count[theme] += 1
            reaction_totals[theme] += reactions
            pos_totals[theme] += pos
            neg_totals[theme] += neg
            neu_totals[theme] += neu
            if views:
                view_totals[theme] += views
                post_counts[theme] += 1

            if quarter:
                b = quarterly[(theme, quarter)]
                if theme == themes[0]:
                    b["primary"] += 1
                b["total"] += 1
                b["reactions"] += reactions
                b["pos"] += pos
                b["neg"] += neg
                b["neu"] += neu
                if views:
                    b["views"] += views
                    b["view_posts"] += 1

        for a, b in combinations(sorted(set(themes)), 2):
            cooccurrence[(a, b)] += 1

    all_themes = sorted(any_count.keys())
    all_quarters = sorted(quarter_totals.keys())
    total_classified = sum(1 for p in posts if p.get("themes"))

    return {
        "themes": all_themes,
        "quarters": all_quarters,
        "total_classified": total_classified,
        "quarter_totals": dict(quarter_totals),
        "primary_count": dict(primary_count),
        "any_count": dict(any_count),
        "reaction_totals": dict(reaction_totals),
        "pos_totals": dict(pos_totals),
        "neg_totals": dict(neg_totals),
        "neu_totals": dict(neu_totals),
        "view_totals": dict(view_totals),
        "post_counts": dict(post_counts),
        "cooccurrence": dict(cooccurrence),
        "quarterly": {f"{t}|{q}": v for (t, q), v in quarterly.items()},
    }


def write_statistics_sheet(wb: Workbook, stats: dict) -> None:
    themes = stats["themes"]
    bold = Font(bold=True)

    # --- Engagement sheet ---
    ws_eng = wb.create_sheet("Theme Stats")
    headers = [
        "theme", "primary_posts", "primary_%", "total_posts", "total_%",
        "avg_reactions", "avg_positive", "avg_negative", "avg_views",
    ]
    ws_eng.append(headers)
    for cell in ws_eng[1]:
        cell.font = bold

    primary = stats["primary_count"]
    any_c = stats["any_count"]
    rxn = stats["reaction_totals"]
    pos = stats["pos_totals"]
    neg = stats["neg_totals"]
    views = stats["view_totals"]
    vcounts = stats["post_counts"]
    total_posts = stats["total_classified"]

    rows = []
    for theme in themes:
        n = any_c.get(theme, 0)
        p = primary.get(theme, 0)
        avg_rxn = round(rxn.get(theme, 0) / n, 1) if n else 0
        avg_pos = round(pos.get(theme, 0) / n, 1) if n else 0
        avg_neg = round(neg.get(theme, 0) / n, 1) if n else 0
        vn = vcounts.get(theme, 0)
        avg_views = round(views.get(theme, 0) / vn, 0) if vn else ""
        primary_pct = round(p / total_posts * 100, 1) if total_posts else 0
        total_pct = round(n / total_posts * 100, 1) if total_posts else 0
        rows.append((theme, p, primary_pct, n, total_pct, avg_rxn, avg_pos, avg_neg, avg_views))

    rows.sort(key=lambda r: r[0])
    for row in rows:
        ws_eng.append(list(row))

    ws_eng.column_dimensions["A"].width = 40
    for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws_eng.column_dimensions[col].width = 16

    # --- Co-occurrence sheet ---
    ws_co = wb.create_sheet("Co-occurrence")
    cooc = stats["cooccurrence"]

    # Header row: theme names
    ws_co.append([""] + themes)
    ws_co[1][0].font = bold
    for cell in ws_co[1][1:]:
        cell.font = bold

    for i, row_theme in enumerate(themes, start=2):
        row = [row_theme]
        ws_co[i][0].font = bold
        for col_theme in themes:
            if row_theme == col_theme:
                row.append("")
            else:
                key = tuple(sorted([row_theme, col_theme]))
                row.append(cooc.get(key, 0) or "")
        ws_co.append(row)

    ws_co.column_dimensions["A"].width = 40
    for i in range(2, len(themes) + 2):
        col_letter = ws_co.cell(row=1, column=i).column_letter
        ws_co.column_dimensions[col_letter].width = 12

    # --- Dynamics (long format) ---
    quarters = stats["quarters"]
    quarter_totals = stats["quarter_totals"]
    qdata = stats["quarterly"]

    ws_dyn = wb.create_sheet("Dynamics")
    dyn_headers = [
        "theme", "quarter",
        "primary_posts", "primary_%", "total_posts", "total_%",
        "avg_reactions", "avg_positive", "avg_negative", "avg_views",
    ]
    ws_dyn.append(dyn_headers)
    for cell in ws_dyn[1]:
        cell.font = bold

    for theme in themes:
        for quarter in quarters:
            b = qdata.get(f"{theme}|{quarter}")
            if not b:
                continue
            n = b["total"]
            p = b["primary"]
            qt = quarter_totals.get(quarter, 0)
            avg_rxn = round(b["reactions"] / n, 1) if n else 0
            avg_pos = round(b["pos"] / n, 1) if n else 0
            avg_neg = round(b["neg"] / n, 1) if n else 0
            avg_views = round(b["views"] / b["view_posts"], 0) if b["view_posts"] else ""
            primary_pct = round(p / qt * 100, 1) if qt else 0
            total_pct = round(n / qt * 100, 1) if qt else 0
            ws_dyn.append([theme, quarter, p, primary_pct, n, total_pct,
                           avg_rxn, avg_pos, avg_neg, avg_views])

    ws_dyn.column_dimensions["A"].width = 40
    ws_dyn.column_dimensions["B"].width = 10
    for col in ["C", "D", "E", "F", "G", "H", "I", "J"]:
        ws_dyn.column_dimensions[col].width = 14

    # --- Dynamics Pivot (total_posts per theme × quarter) ---
    ws_piv = wb.create_sheet("Dynamics Pivot")
    ws_piv.append(["theme"] + quarters)
    for cell in ws_piv[1]:
        cell.font = bold

    for theme in themes:
        row = [theme]
        for quarter in quarters:
            b = qdata.get(f"{theme}|{quarter}")
            row.append(b["total"] if b else "")
        ws_piv.append(row)

    ws_piv.column_dimensions["A"].width = 40
    for i in range(2, len(quarters) + 2):
        col_letter = ws_piv.cell(row=1, column=i).column_letter
        ws_piv.column_dimensions[col_letter].width = 10

    # Per-row color scale: white (min) → blue (max), each row independent
    last_col = ws_piv.cell(row=1, column=len(quarters) + 1).column_letter
    for row_idx in range(2, len(themes) + 2):
        cell_range = f"B{row_idx}:{last_col}{row_idx}"
        ws_piv.conditional_formatting.add(
            cell_range,
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )


def export_posts_to_xlsx(
    channel_dir: Path,
    output_path: Path,
    classifier: BaseClassifier | None = None,
    till: datetime | None = None,
):
    """Export posts from a channel to xlsx file."""
    posts_file = channel_dir / "posts" / "posts.json"
    channel_info_file = channel_dir / "channel_info.json"

    if not posts_file.exists():
        raise FileNotFoundError(f"Posts file not found: {posts_file}")

    with open(posts_file, "r", encoding="utf-8") as f:
        posts = json.load(f)

    if till:
        posts = [p for p in posts if p.get("date") and datetime.fromisoformat(p["date"]).date() <= till]

    # Load channel username for post links
    channel_username = ""
    if channel_info_file.exists():
        with open(channel_info_file, "r", encoding="utf-8") as f:
            channel_info = json.load(f)
            channel_username = channel_info.get("username", "")

    # Classify posts if classifier is provided
    if classifier:
        print("Classifying posts...")
        for post in tqdm(posts, desc="Classifying", unit="post"):
            message = post.get("message", "") or ""
            if message.strip():
                post["themes"] = classifier.classify(message)
            else:
                post["themes"] = []

    wb = Workbook()
    ws = wb.active
    ws.title = "Posts"

    # Header row
    headers = ["post_id", "link", "media_type", "group_id", "date", "time",
               "views", "reactions", "theme_1", "theme_2", "theme_3", "theme_4", "theme_5", "message"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows
    for post in posts:
        post_id = post.get("id", "")
        link = f"https://t.me/{channel_username}/{post_id}" if channel_username else ""
        media_type = get_media_type(post.get("media"))
        group_id = post.get("grouped_id", "") or ""
        date_str, time_str = parse_datetime(post.get("date"))
        message = post.get("message", "") or ""
        views = post.get("views") or ""
        reactions = format_reactions(post.get("reactions"))
        themes = post.get("themes", [])
        theme_cols = [themes[i] if i < len(themes) else "" for i in range(5)]

        ws.append([post_id, link, media_type, group_id, date_str, time_str, views, reactions,
                   *theme_cols, message])

    # Adjust column widths
    ws.column_dimensions["A"].width = 12  # post_id
    ws.column_dimensions["B"].width = 35  # link
    ws.column_dimensions["C"].width = 12  # media_type
    ws.column_dimensions["D"].width = 15  # group_id
    ws.column_dimensions["D"].number_format = numbers.FORMAT_TEXT
    ws.column_dimensions["E"].width = 12  # date
    ws.column_dimensions["F"].width = 10  # time
    ws.column_dimensions["G"].width = 10  # views
    ws.column_dimensions["H"].width = 40  # reactions
    ws.column_dimensions["I"].width = 30  # theme_1
    ws.column_dimensions["J"].width = 30  # theme_2
    ws.column_dimensions["K"].width = 30  # theme_3
    ws.column_dimensions["L"].width = 30  # theme_4
    ws.column_dimensions["M"].width = 30  # theme_5
    ws.column_dimensions["N"].width = 80  # message

    # Statistics sheets (only when themes are present)
    if any(post.get("themes") for post in posts):
        stats = compute_statistics(posts)
        write_statistics_sheet(wb, stats)

    wb.save(output_path)
    print(f"Exported {len(posts)} posts to {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Export channel posts to xlsx file")
    parser.add_argument("channel", help="Channel name (directory name in data/)")
    parser.add_argument(
        "--output", "-o", type=str, default=None,
        help="Output xlsx file path (default: data/<channel>/<channel>.xlsx)"
    )
    parser.add_argument(
        "--till", "-t", type=str, default="2025-12-31",
        help="Ignore posts older than this date (YYYY-MM-DD)"
    )
    parser.add_argument(
        "--classify", "-c",
        type=str,
        default="llm",
        choices=["none", "llm", "keywords", "markers"],
        help="Classification method: none (skip), llm (default), keywords (YAKE extraction), markers (markers.json)"
    )
    args = parser.parse_args()

    channel_dir = Path("data") / args.channel
    if not channel_dir.exists():
        print(f"Error: Channel directory not found: {channel_dir}")
        return 1

    output_path = Path(args.output) if args.output else channel_dir / f"{args.channel}.xlsx"

    till = None
    if args.till:
        till = datetime.fromisoformat(args.till).date()

    classifier = None
    if args.classify != "none":
        classifier = get_classifier(args.classify)

    export_posts_to_xlsx(channel_dir, output_path, classifier, till)
    return 0


if __name__ == "__main__":
    exit(main())